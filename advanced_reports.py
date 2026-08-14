"""advanced_reports.py — Advanced customizable reporting and export functionality.

Allows admins to:
- Select which columns to include in reports
- Choose export format (CSV, Excel, PDF)
- Filter by month, category, storage location
- Generate and preview reports
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import customtkinter as ctk
import csv
import json
from datetime import datetime

from theme import (
    BG_PRIMARY, BG_SURFACE, BG_ELEVATED, BG_HOVER,
    ACCENT, ACCENT_GOLD, ACCENT_GREEN, ACCENT_RED,
    TEXT_PRIMARY, TEXT_SECONDARY, TEXT_MUTED,
    FONT_FAMILY, BORDER_COLOR,
)


class AdvancedReportsDialog(ctk.CTkToplevel):
    """Dialog for generating customizable reports."""

    def __init__(self, parent, db, user: dict = None):
        super().__init__(parent)
        self.db = db
        self.user = user or {}
        self.title("Advanced Reports")
        self.geometry("700x800")
        self.resizable(True, True)
        self.grab_set()
        self._build()
        self.after(100, self.lift)

    def _build(self):
        """Build the reports dialog."""
        # Title
        ctk.CTkLabel(
            self, text="Generate Custom Report",
            font=ctk.CTkFont(size=18, weight="bold"),
        ).pack(pady=(20, 10))

        # Main scrollable frame
        scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # Month selector
        ctk.CTkLabel(
            scroll, text="Month:",
            font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(anchor="w", pady=(10, 5))

        self.month_var = tk.StringVar(value=self.db.get_current_month_year())
        months = self.db.get_all_months()
        if not months:
            months = [self.db.get_current_month_year()]
        
        month_menu = ctk.CTkOptionMenu(scroll, variable=self.month_var, values=months)
        month_menu.pack(fill="x", pady=(0, 15))

        # Column selection
        ctk.CTkLabel(
            scroll, text="Columns to Include:",
            font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(anchor="w", pady=(10, 5))

        self.column_vars = {}
        columns = [
            ("item_name", "Item Name"),
            ("category", "Category"),
            ("storage_location", "Storage Location"),
            ("current_lbs", "Current Pounds"),
            ("donated_lbs", "Donated Pounds"),
            ("discarded_lbs", "Discarded Pounds"),
            ("remaining_lbs", "Remaining Pounds"),
            ("current_qty", "Current Quantity"),
            ("recorded_by", "Recorded By"),
        ]

        for col_key, col_label in columns:
            var = tk.BooleanVar(value=True)
            self.column_vars[col_key] = var
            ctk.CTkCheckBox(
                scroll, text=col_label, variable=var,
                font=ctk.CTkFont(size=11),
            ).pack(anchor="w", pady=4)

        # Export format
        ctk.CTkLabel(
            scroll, text="Export Format:",
            font=ctk.CTkFont(size=12, weight="bold"),
        ).pack(anchor="w", pady=(15, 5))

        self.format_var = tk.StringVar(value="CSV")
        for fmt in ["CSV", "Excel", "PDF"]:
            ctk.CTkRadioButton(
                scroll, text=fmt, variable=self.format_var, value=fmt,
                font=ctk.CTkFont(size=11),
            ).pack(anchor="w", pady=4)

        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(0, 20))
        btn_frame.grid_columnconfigure(0, weight=1)

        ctk.CTkButton(
            btn_frame, text="Preview", width=140, height=40,
            fg_color=ACCENT, hover_color="#FF9500",
            text_color="white",
            command=self._preview_report,
        ).pack(side="right", padx=(10, 0))

        ctk.CTkButton(
            btn_frame, text="Export", width=140, height=40,
            fg_color=ACCENT_GREEN, hover_color="#16a34a",
            text_color="white",
            command=self._export_report,
        ).pack(side="right", padx=(0, 10))

        ctk.CTkButton(
            btn_frame, text="Cancel", width=110, height=40,
            fg_color="#7f8c8d", hover_color="#626567",
            command=self.destroy,
        ).pack(side="right")

    def _get_selected_columns(self) -> list:
        """Get list of selected columns."""
        return [col for col, var in self.column_vars.items() if var.get()]

    def _generate_report_data(self) -> list:
        """Generate report data based on selections."""
        month = self.month_var.get()
        selected_cols = self._get_selected_columns()

        try:
            # Get weight data
            conn = self.db._connect()
            cursor = conn.cursor()
            cursor.execute("""
                SELECT i.id, i.item_name, i.category, i.storage_location,
                       i.current_quantity, wh.current_pounds, wh.donated_pounds,
                       wh.discarded_pounds, wh.calculated_remaining, wh.recorded_by
                FROM weight_history wh
                JOIN inventory_items i ON wh.item_id = i.id
                WHERE wh.month_year = ?
                ORDER BY i.item_name
            """, (month,))

            rows = cursor.fetchall()
            conn.close()
            data = []

            for row in rows:
                item_id, name, category, location, qty, curr, donated, discarded, remaining, recorded_by = row
                item_data = {
                    "item_name": name,
                    "category": category or "—",
                    "storage_location": location or "—",
                    "current_qty": qty,
                    "current_lbs": f"{curr:.2f}",
                    "donated_lbs": f"{donated:.2f}",
                    "discarded_lbs": f"{discarded:.2f}",
                    "remaining_lbs": f"{remaining:.2f}",
                    "recorded_by": recorded_by or "—",
                }
                data.append(item_data)

            return data
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
            return []

    def _preview_report(self):
        """Preview the report in a new window."""
        data = self._generate_report_data()
        if not data:
            messagebox.showwarning("No Data", "No data available for this month")
            return

        selected_cols = self._get_selected_columns()
        month = self.month_var.get()

        # Create preview window
        preview = ctk.CTkToplevel(self)
        preview.title(f"Report Preview - {month}")
        preview.geometry("900x600")

        # Title
        ctk.CTkLabel(
            preview, text=f"Monthly Report - {month}",
            font=ctk.CTkFont(size=16, weight="bold"),
        ).pack(pady=(10, 5))

        # Summary
        summary = self.db.get_weight_summary(month)
        summary_text = (
            f"Total Current: {summary['total_current']:.2f} lbs | "
            f"Total Donated: {summary['total_donated']:.2f} lbs | "
            f"Total Discarded: {summary['total_discarded']:.2f} lbs | "
            f"Total Remaining: {summary['total_remaining']:.2f} lbs"
        )
        ctk.CTkLabel(
            preview, text=summary_text,
            font=ctk.CTkFont(size=10),
            text_color=TEXT_MUTED,
        ).pack(pady=(0, 10))

        # Text preview
        text_frame = ctk.CTkFrame(preview)
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)

        text_widget = ctk.CTkTextbox(text_frame)
        text_widget.pack(fill="both", expand=True)

        # Build preview text
        preview_lines = []
        preview_lines.append(f"Monthly Report - {month}")
        preview_lines.append("=" * 100)
        preview_lines.append("")

        # Header
        header = " | ".join([col.replace("_", " ").title() for col in selected_cols])
        preview_lines.append(header)
        preview_lines.append("-" * 100)

        # Data rows
        for item in data:
            row = " | ".join([str(item.get(col, "—")) for col in selected_cols])
            preview_lines.append(row)

        preview_lines.append("")
        preview_lines.append(summary_text)

        text_widget.insert("1.0", "\n".join(preview_lines))
        text_widget.configure(state="disabled")

    def _export_report(self):
        """Export the report to selected format."""
        data = self._generate_report_data()
        if not data:
            messagebox.showwarning("No Data", "No data available for this month")
            return

        month = self.month_var.get()
        fmt = self.format_var.get()
        selected_cols = self._get_selected_columns()

        # Sanitize month for filename
        filename_month = month.replace(" ", "_")

        if fmt == "CSV":
            self._export_csv(data, selected_cols, filename_month, month)
        elif fmt == "Excel":
            self._export_excel(data, selected_cols, filename_month, month)
        elif fmt == "PDF":
            self._export_pdf(data, selected_cols, filename_month, month)

    def _export_csv(self, data, columns, filename_month, month):
        """Export report as CSV."""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"Weights_Report_{filename_month}.csv"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Header
                writer.writerow([f"Monthly Report - {month}"])
                writer.writerow([])

                # Column headers
                headers = [col.replace("_", " ").title() for col in columns]
                writer.writerow(headers)

                # Data rows
                for item in data:
                    row = [str(item.get(col, "—")) for col in columns]
                    writer.writerow(row)

                # Summary
                writer.writerow([])
                summary = self.db.get_weight_summary(month)
                writer.writerow(["Summary"])
                writer.writerow(["Total Current Pounds", f"{summary['total_current']:.2f}"])
                writer.writerow(["Total Donated Pounds", f"{summary['total_donated']:.2f}"])
                writer.writerow(["Total Discarded Pounds", f"{summary['total_discarded']:.2f}"])
                writer.writerow(["Total Remaining Pounds", f"{summary['total_remaining']:.2f}"])

            messagebox.showinfo("Success", f"Report exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")

    def _export_excel(self, data, columns, filename_month, month):
        """Export report as Excel (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Error", "openpyxl not installed. Install with: pip install openpyxl")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Weights_Report_{filename_month}.xlsx"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Weights"

            # Title
            ws['A1'] = f"Monthly Report - {month}"
            ws['A1'].font = Font(bold=True, size=14)

            # Headers
            headers = [col.replace("_", " ").title() for col in columns]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            # Data
            for row_idx, item in enumerate(data, 4):
                for col_idx, col in enumerate(columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = item.get(col, "—")
                    cell.alignment = Alignment(horizontal="center")

            # Summary
            summary_row = len(data) + 5
            ws[f'A{summary_row}'] = "Summary"
            ws[f'A{summary_row}'].font = Font(bold=True)

            summary = self.db.get_weight_summary(month)
            ws[f'A{summary_row + 1}'] = "Total Current Pounds"
            ws[f'B{summary_row + 1}'] = summary['total_current']
            ws[f'A{summary_row + 2}'] = "Total Donated Pounds"
            ws[f'B{summary_row + 2}'] = summary['total_donated']
            ws[f'A{summary_row + 3}'] = "Total Discarded Pounds"
            ws[f'B{summary_row + 3}'] = summary['total_discarded']
            ws[f'A{summary_row + 4}'] = "Total Remaining Pounds"
            ws[f'B{summary_row + 4}'] = summary['total_remaining']

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

            wb.save(file_path)
            messagebox.showinfo("Success", f"Report exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")

    def _export_pdf(self, data, columns, filename_month, month):
        """Export report as PDF (requires reportlab)."""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
        except ImportError:
            messagebox.showerror("Error", "reportlab not installed. Install with: pip install reportlab")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile=f"Weights_Report_{filename_month}.pdf"
        )

        if not file_path:
            return

        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f2937'),
                spaceAfter=12,
            )
            elements.append(Paragraph(f"Monthly Report - {month}", title_style))
            elements.append(Spacer(1, 0.2*inch))

            # Summary
            summary = self.db.get_weight_summary(month)
            summary_text = (
                f"Total Current: {summary['total_current']:.2f} lbs | "
                f"Total Donated: {summary['total_donated']:.2f} lbs | "
                f"Total Discarded: {summary['total_discarded']:.2f} lbs | "
                f"Total Remaining: {summary['total_remaining']:.2f} lbs"
            )
            elements.append(Paragraph(summary_text, styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))

            # Table
            headers = [col.replace("_", " ").title() for col in columns]
            table_data = [headers]
            for item in data:
                row = [str(item.get(col, "—")) for col in columns]
                table_data.append(row)

            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)

            doc.build(elements)
            messagebox.showinfo("Success", f"Report exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")
